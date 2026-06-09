import os
import re
import json
import uuid
import time
import random
import hashlib
import threading
from datetime import datetime
from typing import Dict, Any, List

from bs4 import BeautifulSoup
from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient
from openai import AzureOpenAI
from openpyxl import Workbook, load_workbook

from dotenv import load_dotenv
load_dotenv()

# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------

CREDIT_REPORT_SCHEMA = {
    # Report Meta
    "report_type": None,
    "report_date": None,
    "requested_by": None,
    "property": None,
    "application_id": None,

    # Applicant Identity
    "full_name": None,
    "dob": None,
    "ssn": None,
    "aka": None,
    "aka_count": None,
    "current_address": None,
    "previous_address": None,
    "second_previous_address": None,
    "phone": None,
    "employment": None,

    # SSN / Fraud Indicators
    "ssn_message": None,
    "fraud_indicators": None,
    "ssn_alert_code": None,
    "address_conflict": None,
    "fraud_indicator_count": None,

    # Credit Score
    "credit_score": None,
    "credit_score_model": None,
    "credit_score_source": None,
    "score_factors": [],

    # Record Counts
    "tradelines_count": None,
    "collections_count": None,
    "public_records_count": None,
    "inquiries_count": None,
    "rapid_tradeline_flag": None,

    # Derogatory Items
    "negative_tradelines": None,
    "tradelines_with_historical_negatives": None,
    "occurrence_of_historical_negatives": None,

    # Tradeline Summary (JSON)
    "tradeline_summary": {},

    # Late Payment Totals
    "total_late_30": None,
    "total_late_60": None,
    "total_late_90": None,
    "total_late_120_plus": None,

    # Arrays (stored as JSON in raw sheet)
    "tradelines": [],
    "collections": [],
    "inquiries": [],

    # HTML Screening Extras
    "decision": None,
    "decision_notes": None,
    "criteria_results": [],
    "criminal_records_any": None,
    "eviction_records_any": None,
    "sanctions_check": None,

    # Income (HTML Screening Reports only)
    "monthly_rent": None,
    "monthly_income": None,
    "rent_to_income_pct": None,
    "min_income_required": None,
}

TRADELINE_ITEM = {
    "creditor_name": None, "industry": None, "loan_type": None,
    "loan_terms": None, "account_type": None, "opened": None,
    "verified": None, "closed": None, "balance": None,
    "past_due": None, "credit_limit": None, "payment_amount": None,
    "status": None, "remarks": None, "notes": None,
    "late_30": None, "late_60": None, "late_90": None, "late_120_plus": None,
}

COLLECTION_ITEM = {
    "collector_name": None, "industry": None, "opened": None,
    "verified": None, "current_balance": None, "original_balance": None,
    "past_due": None, "status": None, "remarks": None, "credit_grantor": None,
}

INQUIRY_ITEM = {
    "date": None, "name": None, "kind_of_business": None,
}

# Tradeline summary types
TRADELINE_SUMMARY_TYPES = ["revolving", "installment", "mortgage", "open", "closed_w_bal", "total"]
TRADELINE_SUMMARY_FIELDS = ["count", "high_credit", "credit_limit", "balance", "past_due", "payment", "available"]

# HTML criteria codes to flatten (most common)
CRITERIA_CODES = [
    "CR100", "CR101", "CR112", "CR126", "CR127", "CR128", "CR131", "CR132",
    "GS306",
    "CM321", "CM322", "CM323", "CM325", "CM326", "CM327", "CM328",
    "CM329", "CM330", "CM331", "CM332", "CM333", "CM334", "CM335",
    "CM337", "CM338", "CM339", "CM340",
    "EV701", "EV702",
    "CO807", "CO808",
    "SW849", "SW851", "SW852", "SW853",
    "SN901", "SN902",
    "SX501", "RI400",
]


# ---------------------------------------------------------------------------
# Flat analytics row builder
# ---------------------------------------------------------------------------

def build_flat_row(data: Dict[str, Any], filename: str) -> Dict[str, Any]:
    """Expand all JSON fields into individual columns for data analysis / ML."""
    flat = {}

    # --- Identity & meta (direct copy) ---
    direct_fields = [
        "report_type", "report_date", "requested_by", "property", "application_id",
        "full_name", "dob", "ssn", "aka", "aka_count",
        "current_address", "previous_address", "second_previous_address",
        "phone", "employment",
        "ssn_message", "fraud_indicators", "ssn_alert_code", "address_conflict", "fraud_indicator_count",
        "credit_score", "credit_score_model", "credit_score_source",
        "tradelines_count", "collections_count", "public_records_count",
        "inquiries_count", "rapid_tradeline_flag",
        "negative_tradelines", "tradelines_with_historical_negatives",
        "occurrence_of_historical_negatives",
        "total_late_30", "total_late_60", "total_late_90", "total_late_120_plus",
        "decision", "decision_notes", "criminal_records_any",
        "eviction_records_any", "sanctions_check",
        "monthly_rent", "monthly_income", "rent_to_income_pct", "min_income_required",
    ]
    for f in direct_fields:
        flat[f] = data.get(f)

    # --- Fraud indicator count ---
    fraud_indicators_val = data.get("fraud_indicators")
    if fraud_indicators_val and isinstance(fraud_indicators_val, str) and fraud_indicators_val.strip():
        flat["fraud_indicator_count"] = fraud_indicators_val.count("|") + 1
    else:
        flat["fraud_indicator_count"] = 0

    # --- Score factors → score_factor_1 .. score_factor_4 ---
    score_factors = data.get("score_factors") or []
    if isinstance(score_factors, str):
        try:
            score_factors = json.loads(score_factors)
        except Exception:
            score_factors = []
    for i in range(1, 5):
        flat[f"score_factor_{i}"] = score_factors[i - 1] if i <= len(score_factors) else None

    # --- Tradeline summary → tradeline_summary_<type>_<field> ---
    summary = data.get("tradeline_summary") or {}
    if isinstance(summary, str):
        try:
            summary = json.loads(summary)
        except Exception:
            summary = {}
    for t in TRADELINE_SUMMARY_TYPES:
        t_data = summary.get(t) or {}
        for f in TRADELINE_SUMMARY_FIELDS:
            flat[f"tradeline_summary_{t}_{f}"] = t_data.get(f)

    # --- Tradelines → aggregated stats ---
    tradelines = data.get("tradelines") or []
    if isinstance(tradelines, str):
        try:
            tradelines = json.loads(tradelines)
        except Exception:
            tradelines = []

    balances      = [t.get("balance")        for t in tradelines if t.get("balance")        is not None]
    past_dues     = [t.get("past_due")       for t in tradelines if t.get("past_due")       is not None]
    credit_limits = [t.get("credit_limit")   for t in tradelines if t.get("credit_limit")   is not None]
    payments      = [t.get("payment_amount") for t in tradelines if t.get("payment_amount") is not None]

    flat["tradeline_total_balance"]        = sum(balances)      if balances      else None
    flat["tradeline_total_past_due"]       = sum(past_dues)     if past_dues     else None
    flat["tradeline_total_credit_limit"]   = sum(credit_limits) if credit_limits else None
    flat["tradeline_total_payment_amount"] = sum(payments)      if payments      else None
    flat["tradeline_max_balance"]          = max(balances)      if balances      else None
    flat["tradeline_max_past_due"]         = max(past_dues)     if past_dues     else None

    charged_off   = [t for t in tradelines if "charged off"  in str(t.get("status", "")).lower()]
    collections_t = [t for t in tradelines if "collection"   in str(t.get("status", "")).lower()]
    derogatory    = [t for t in tradelines if any(
        kw in str(t.get("status", "")).lower()
        for kw in ["charged off", "collection", "past due", "derogatory"]
    )]
    open_accounts = [t for t in tradelines if not t.get("closed")]

    flat["tradeline_charged_off_count"]     = len(charged_off)
    flat["tradeline_collection_count"]      = len(collections_t)
    flat["tradeline_derogatory_count"]      = len(derogatory)
    flat["tradeline_open_account_count"]    = len(open_accounts)

    late_30  = [t.get("late_30",  0) or 0 for t in tradelines]
    late_60  = [t.get("late_60",  0) or 0 for t in tradelines]
    late_90  = [t.get("late_90",  0) or 0 for t in tradelines]
    late_120 = [t.get("late_120_plus", 0) or 0 for t in tradelines]

    flat["tradeline_max_late_30_days"]      = max(late_30)  if late_30  else None
    flat["tradeline_max_late_60_days"]      = max(late_60)  if late_60  else None
    flat["tradeline_max_late_90_days"]      = max(late_90)  if late_90  else None
    flat["tradeline_max_late_120_plus_days"]= max(late_120) if late_120 else None
    flat["tradeline_accounts_with_any_late"]= sum(
        1 for t in tradelines
        if any((t.get(k) or 0) > 0 for k in ["late_30","late_60","late_90","late_120_plus"])
    )

    # --- Collections aggregated ---
    collections = data.get("collections") or []
    if isinstance(collections, str):
        try:
            collections = json.loads(collections)
        except Exception:
            collections = []

    col_balances  = [c.get("current_balance")  for c in collections if c.get("current_balance")  is not None]
    col_originals = [c.get("original_balance") for c in collections if c.get("original_balance") is not None]
    col_past_dues = [c.get("past_due")         for c in collections if c.get("past_due")         is not None]

    flat["collection_total_current_balance"]  = sum(col_balances)  if col_balances  else None
    flat["collection_total_original_balance"] = sum(col_originals) if col_originals else None
    flat["collection_total_past_due"]         = sum(col_past_dues) if col_past_dues else None
    flat["collection_max_balance"]            = max(col_balances)  if col_balances  else None
    flat["collection_unique_creditor_count"]  = len({c.get("credit_grantor") for c in collections if c.get("credit_grantor")})

    # --- Inquiries aggregated ---
    inquiries = data.get("inquiries") or []
    if isinstance(inquiries, str):
        try:
            inquiries = json.loads(inquiries)
        except Exception:
            inquiries = []

    flat["inquiry_total_count"]           = len(inquiries)
    flat["inquiry_unique_creditor_count"] = len({i.get("name") for i in inquiries if i.get("name")})
    flat["inquiry_most_recent_date"]      = inquiries[0].get("date") if inquiries else None

    # --- Criteria results → one column per code (HTML reports) ---
    criteria = data.get("criteria_results") or []
    if isinstance(criteria, str):
        try:
            criteria = json.loads(criteria)
        except Exception:
            criteria = []

    criteria_map = {c.get("code"): c.get("result") for c in criteria if c.get("code")}
    for code in CRITERIA_CODES:
        flat[f"criteria_{code}"] = criteria_map.get(code)

    flat["source_filename"] = filename
    flat["uploaded_at"] = datetime.now().strftime("%m/%d/%Y %I:%M:%S %p")

    return flat


# ---------------------------------------------------------------------------
# Text extraction
# ---------------------------------------------------------------------------

def extract_text_with_ocr(file_content: bytes) -> str:
    endpoint = os.getenv("AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT")
    key = os.getenv("AZURE_DOCUMENT_INTELLIGENCE_KEY")
    if not endpoint or not key:
        raise ValueError("Azure Document Intelligence credentials not configured")
    client = DocumentIntelligenceClient(
        endpoint=endpoint,
        credential=AzureKeyCredential(key),
        api_version="2024-11-30",
    )
    poller = client.begin_analyze_document(
        model_id="prebuilt-read",
        body=file_content,
        content_type="application/octet-stream",
    )
    result = poller.result()
    if not result or not hasattr(result, "content"):
        raise ValueError("No text extracted from document")
    return result.content


def extract_text_from_html(file_content: bytes) -> str:
    soup = BeautifulSoup(file_content.decode("utf-8", errors="ignore"), "html.parser")
    return soup.get_text(separator="\n", strip=True)


# ---------------------------------------------------------------------------
# GPT extraction
# ---------------------------------------------------------------------------

def extract_with_gpt(text: str) -> Dict[str, Any]:
    endpoint   = os.getenv("AZURE_OPENAI_ENDPOINT")
    key        = os.getenv("AZURE_OPENAI_API_KEY")
    deployment = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME")
    api_version= os.getenv("AZURE_OPENAI_API_VERSION", "2025-04-01-preview")
    if not endpoint or not key:
        raise ValueError("Azure OpenAI credentials not configured")

    client = AzureOpenAI(azure_endpoint=endpoint, api_key=key, api_version=api_version)

    extraction_id = str(uuid.uuid4())
    random_seed   = random.randint(100000, 999999)
    doc_hash      = hashlib.sha256(text.encode()).hexdigest()[:16]
    time.sleep(0.5)

    schema_json   = json.dumps(CREDIT_REPORT_SCHEMA, indent=2)
    tradeline_json= json.dumps(TRADELINE_ITEM)
    collection_json=json.dumps(COLLECTION_ITEM)
    inquiry_json  = json.dumps(INQUIRY_ITEM)

    prompt = f"""Extract credit report data and return JSON matching EXACTLY this schema (null for missing values, [] for missing lists):

{schema_json}

Each item in `tradelines` must follow: {tradeline_json}
Each item in `collections` must follow: {collection_json}
Each item in `inquiries` must follow: {inquiry_json}

Rules:
- `report_type`: "TransUnion PDF" or "HTML Screening Report".
- `aka_count`: count distinct AKA aliases (>2 may indicate identity fraud).
- `rapid_tradeline_flag`: true if 3+ tradelines opened within any 6-month window.
- `address_conflict`: true if report contains an address mismatch warning.
- `fraud_indicator_count`: count of distinct fraud/alert descriptions found. 0 if none.
- `fraud_indicators`: concatenate ALL fraud/alert descriptions separated by " | ". null if none.
- `credit_score`: integer or null if not scored / insufficient credit.
- `score_factors`: list of key factor strings.
- `tradeline_summary`: keys revolving, installment, mortgage, open, closed_w_bal, total. Each: {{ count, high_credit, credit_limit, balance, past_due, payment, available }} — null for N/A.
- `total_late_30/60/90/120_plus`: sum across ALL tradelines.
- `negative_tradelines`, `tradelines_with_historical_negatives`, `occurrence_of_historical_negatives`: integers.
- `decision` (HTML only): "Approved", "Approved with Conditions", or "Declined".
- `criteria_results` (HTML only): array of {{ code, description, result }} where result is "P","F","*","N","--".
- `criminal_records_any`: true if ANY CM-prefixed criteria result = "F".
- `eviction_records_any`: true if ANY EV-prefixed criteria result = "F".
- `sanctions_check`: "Pass" if GS306="P", "Fail" if "F", null if absent.
- `monthly_rent`: numeric monthly rent from the report header (e.g. "Monthly Rent: $920" → 920). null if not present.
- `monthly_income`: numeric monthly income from the Rent to Income Summary table (e.g. "$ 4,500" → 4500). null if not present.
- `rent_to_income_pct`: numeric percentage from Rent to Income Summary (e.g. "20.44 %" → 20.44). null if not present.
- `min_income_required`: numeric minimum income required from Rent to Income Summary (e.g. "$ 2,629" → 2629). null if not present.
- monetary values as numbers (strip $ and commas), null if N/A.
- dates as MM/YY or MM/DD/YYYY as they appear.
- Return ONLY valid JSON. No markdown, no commentary.

Credit report text:
{text}
"""

    response = client.chat.completions.create(
        model=deployment,
        messages=[
            {
                "role": "system",
                "content": (
                    f"SESSION: {extraction_id} HASH: {doc_hash} "
                    "You are a precise credit report data extractor for a property management company. "
                    "Return only valid JSON."
                ),
            },
            {"role": "user", "content": prompt},
        ],
        temperature=0.1,
        max_completion_tokens=8000,
        seed=random_seed,
        user=f"extraction_{extraction_id}",
        response_format={"type": "json_object"},
    )

    result_text = response.choices[0].message.content or ""
    result_text = re.sub(r"^```(?:json)?\s*|\s*```$", "", result_text.strip())

    try:
        result = json.loads(result_text)
    except json.JSONDecodeError:
        cleaned = re.sub(r",\s*([}\]])", r"\1", result_text)
        result = json.loads(cleaned)

    for k, default in CREDIT_REPORT_SCHEMA.items():
        if k not in result or (result[k] is None and isinstance(default, list)):
            result[k] = [] if isinstance(default, list) else (
                {} if isinstance(default, dict) else default
            )
    return result


# ---------------------------------------------------------------------------
# Single Excel, single sheet — raw JSON columns + flat columns side by side
# ---------------------------------------------------------------------------

JSON_FIELDS = ["score_factors", "tradeline_summary", "tradelines",
               "collections", "inquiries", "criteria_results"]


def _build_combined_row(extracted: Dict[str, Any], flat_row: Dict[str, Any],
                        filename: str) -> Dict[str, Any]:
    """
    One dict with every column:
      1. Raw schema fields (JSON arrays/objects stored as JSON strings)
      2. All flat/expanded columns from build_flat_row()
    source_filename and uploaded_at come from flat_row (already set there).
    """
    combined = {}

    # --- Raw JSON side ---
    for key in CREDIT_REPORT_SCHEMA:
        val = extracted.get(key)
        if key in JSON_FIELDS:
            if not isinstance(val, str):
                val = json.dumps(val or ([] if isinstance(val, list) else {}),
                                 ensure_ascii=False)
        combined[key] = val

    # --- Flat/expanded side (prefixed to avoid collisions with raw keys) ---
    # flat_row already contains source_filename and uploaded_at so we skip
    # adding them again from raw side.
    for key, val in flat_row.items():
        if key not in combined:          # don't overwrite raw fields
            combined[key] = val

    return combined


def _safe_cell(val):
    """Coerce any value to something openpyxl can write to a cell."""
    if val is None:
        return ""
    if isinstance(val, (list, dict)):
        import json as _json
        return _json.dumps(val, ensure_ascii=False)
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val
    return str(val)


def append_to_excel(extracted: Dict[str, Any], flat_row: Dict[str, Any],
                    excel_path: str) -> None:
    """Append one combined row to the single Excel file."""
    os.makedirs(os.path.dirname(os.path.abspath(excel_path)), exist_ok=True)

    combined = _build_combined_row(extracted, flat_row,
                                   flat_row.get("source_filename", ""))

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        existing_headers = [ws.cell(1, c).value
                            for c in range(1, ws.max_column + 1)]
        for key in combined:
            if key not in existing_headers:
                existing_headers.append(key)
                ws.cell(1, len(existing_headers)).value = key
        ws.append([_safe_cell(combined.get(h)) for h in existing_headers])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Credit Reports"
        headers = list(combined.keys())
        ws.append(headers)
        ws.append([_safe_cell(combined.get(h)) for h in headers])

    wb.save(excel_path)


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def extract_credit_report_data(
    file_content: bytes,
    filename: str,
    excel_path: str,
) -> Dict[str, Any]:
    ext = os.path.splitext(filename)[1].lower()

    if ext in (".htm", ".html"):
        text = extract_text_from_html(file_content)
    elif ext in (".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp"):
        text = extract_text_with_ocr(file_content)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    if not text or len(text.strip()) < 50:
        raise ValueError("Insufficient text extracted from document")

    extracted = extract_with_gpt(text)
    flat_row  = build_flat_row(extracted, filename)
    append_to_excel(extracted, flat_row, excel_path)
    return flat_row


# ---------------------------------------------------------------------------
# Batch helper — parallel extraction, serialised Excel writes
# ---------------------------------------------------------------------------

def process_batch(
    files: List[tuple],       # list of (file_content_bytes, filename)
    excel_path: str,
    max_workers: int = 4,
) -> List[Dict[str, Any]]:
    import concurrent.futures

    lock = threading.Lock()

    def _process(item):
        content, fname = item
        try:
            ext = os.path.splitext(fname)[1].lower()
            if ext in (".htm", ".html"):
                text = extract_text_from_html(content)
            elif ext in (".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp"):
                text = extract_text_with_ocr(content)
            else:
                raise ValueError(f"Unsupported file type: {ext}")

            if not text or len(text.strip()) < 50:
                raise ValueError("Insufficient text extracted")

            extracted = extract_with_gpt(text)
            flat_row  = build_flat_row(extracted, fname)

            with lock:
                append_to_excel(extracted, flat_row, excel_path)

            return {"filename": fname, "success": True, "data": flat_row}
        except Exception as e:
            return {"filename": fname, "success": False, "error": str(e)}

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        results = list(executor.map(_process, files))

    return results